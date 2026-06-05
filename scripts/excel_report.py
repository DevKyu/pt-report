"""
excel_report.py — Excel 리포트 생성
process.py의 load_and_process() 결과를 받아 xlsx 생성
"""
from __future__ import annotations
from typing import Optional, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import datetime

# ── 색상 팔레트 ─────────────────────────────────────────────────────────
HDR    = "1E3A5F"; HDR_FG = "FFFFFF"
IN_H   = "2D6A9F"; OUT_H  = "A05C0A"
TOT_H  = "1E6B45"; WK_H   = "3D4F7A"
ROW_W  = "FFFFFF"; ROW_A  = "F2F6FC"
SUB_BG = "E4EAF5"; TOT_BG = "CDD9EE"
RED_FG = "C0392B"; NRM_FG = "1A1A1A"
MUT_FG = "888888"; WK_FG  = "7A8EAA"
HDR_NAME = "1E3A5F"
BORD   = "C5D0E8"

def bs(thick_bottom=False):
    s = Side(style='thin', color=BORD)
    b = Side(style='medium', color="7A98C0") if thick_bottom else s
    return Border(left=s, right=s, top=s, bottom=b)

def hc(ws, r, c, v, bg=HDR, fg=HDR_FG, sz=10, align='center', bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='맑은 고딕', bold=bold, color=fg, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = bs()
    return cell

def dc(ws, r, c, v, bg=ROW_W, fg=NRM_FG, bold=False, sz=10, align='center', thick=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='맑은 고딕', bold=bold, color=fg, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = bs(thick)
    return cell

def title_cell(ws, r, ncols, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name='맑은 고딕', bold=True, size=14, color=HDR)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 36

def add_note(ws, r, ncols, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name='맑은 고딕', size=9, color="999999", italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 16

def generate_excel(data: dict, output_path: Union[str, Path], validation_result: dict = None) -> Path:
    """
    집계 데이터를 받아 Excel 파일 생성
    
    Args:
        data: process.load_and_process() 반환값
        output_path: 저장 경로
    
    Returns:
        생성된 파일 Path
    """
    summary  = data['summary']
    daily    = data['daily']
    weekly   = data['weekly']
    docs     = data['docs']
    weeks    = data['weeks']
    period   = data['period']

    wb = openpyxl.Workbook()

    # ── 시트1: 월별 요약 ───────────────────────────────────────────────
    ws = wb.active; ws.title = "월별 요약"
    ws.sheet_view.showGridLines = False
    title_cell(ws, 1, 4, f"처방의별 월간 집계 — {period}")
    ws.row_dimensions[2].height = 6

    hc(ws, 3, 1, '처방의', align='left')
    hc(ws, 3, 2, '입원', bg=IN_H)
    hc(ws, 3, 3, '외래', bg=OUT_H)
    hc(ws, 3, 4, '합계', bg=TOT_H)
    ws.row_dimensions[3].height = 26
    ws.freeze_panes = 'A4'

    for ri, row in summary.iterrows():
        r = ri + 4
        bg = ROW_A if ri % 2 == 0 else ROW_W
        dc(ws, r, 1, row['처방의'], bg=bg, align='left', bold=True, fg=HDR_NAME)
        dc(ws, r, 2, int(row['입원']), bg=bg)
        dc(ws, r, 3, int(row['외래']), bg=bg)
        dc(ws, r, 4, int(row['합계']), bg=bg, bold=True)
        ws.row_dimensions[r].height = 24

    tr = len(summary) + 4
    hc(ws, tr, 1, '합계', align='left')
    dc(ws, tr, 2, int(summary['입원'].sum()), bg=TOT_BG, bold=True)
    dc(ws, tr, 3, int(summary['외래'].sum()), bg=TOT_BG, bold=True)
    dc(ws, tr, 4, int(summary['합계'].sum()), bg=TOT_BG, bold=True)
    ws.row_dimensions[tr].height = 28

    ws.column_dimensions['A'].width = 13
    for col in ['B','C','D']: ws.column_dimensions[col].width = 13

    # ── 시트2: 주별 집계 ───────────────────────────────────────────────
    ws2 = wb.create_sheet("주별 집계")
    ws2.sheet_view.showGridLines = False
    title_cell(ws2, 1, 5, f"처방의별 주별 집계 — {period}")
    ws2.row_dimensions[2].height = 6

    hc(ws2, 3, 1, '처방의', align='left')
    hc(ws2, 3, 2, '주차',  bg=WK_H)
    hc(ws2, 3, 3, '입원',  bg=IN_H)
    hc(ws2, 3, 4, '외래',  bg=OUT_H)
    hc(ws2, 3, 5, '합계',  bg=TOT_H)
    ws2.row_dimensions[3].height = 26
    ws2.freeze_panes = 'A4'

    prev_doc = None; ri = 4
    for _, row in weekly.iterrows():
        doc = row['처방의']
        bg = ROW_A if ri % 2 == 0 else ROW_W
        dc(ws2, ri, 1, doc if doc != prev_doc else '', bg=bg, align='left', bold=(doc!=prev_doc), fg=HDR_NAME)
        dc(ws2, ri, 2, row['주차'], bg=bg, fg=WK_FG, sz=9)
        dc(ws2, ri, 3, int(row['입원']), bg=bg)
        dc(ws2, ri, 4, int(row['외래']), bg=bg)
        dc(ws2, ri, 5, int(row['합계']), bg=bg, bold=True)
        ws2.row_dimensions[ri].height = 22

        nx = weekly.iloc[weekly.index.get_loc(_)+1:weekly.index.get_loc(_)+2]
        if nx.empty or nx.iloc[0]['처방의'] != doc:
            ri += 1
            d = weekly[weekly['처방의']==doc]
            bg2 = ROW_A if ri % 2 == 0 else ROW_W
            dc(ws2, ri, 1, f'{doc} 소계', bg=bg2, align='left', bold=True, fg=HDR_NAME, thick=True)
            dc(ws2, ri, 2, '', bg=bg2, thick=True)
            dc(ws2, ri, 3, int(d['입원'].sum()), bg=bg2, bold=True, thick=True)
            dc(ws2, ri, 4, int(d['외래'].sum()), bg=bg2, bold=True, thick=True)
            dc(ws2, ri, 5, int(d['합계'].sum()), bg=bg2, bold=True, thick=True)
            ws2.row_dimensions[ri].height = 22
        prev_doc = doc; ri += 1

    hc(ws2, ri, 1, '전체 합계', align='left')
    dc(ws2, ri, 2, '', bg=TOT_BG)
    dc(ws2, ri, 3, int(weekly['입원'].sum()), bg=TOT_BG, bold=True)
    dc(ws2, ri, 4, int(weekly['외래'].sum()), bg=TOT_BG, bold=True)
    dc(ws2, ri, 5, int(weekly['합계'].sum()), bg=TOT_BG, bold=True)
    ws2.row_dimensions[ri].height = 28

    ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 13
    for col in ['C','D','E']: ws2.column_dimensions[col].width = 12

    # ── 시트3: 일별 상세 ───────────────────────────────────────────────
    ws3 = wb.create_sheet("일별 상세")
    ws3.sheet_view.showGridLines = False
    title_cell(ws3, 1, 7, f"처방의별 일별 상세 — {period}")
    add_note(ws3, 2, 7, "날짜 빨간색 = 토요일 · 공휴일  |  주차별 소계는 각 처방의 탭 참조")
    ws3.row_dimensions[3].height = 6

    for ci, (h, bg) in enumerate([
        ('처방의',HDR),('접수일자',HDR),('요일',HDR),
        ('주차',WK_H),('입원',IN_H),('외래',OUT_H),('합계',TOT_H)
    ], 1):
        hc(ws3, 4, ci, h, bg=bg, align='left' if ci<=2 else 'center')
    ws3.row_dimensions[4].height = 26
    ws3.freeze_panes = 'A5'

    prev_doc = None; ri = 5
    sorted_daily = daily.sort_values(['처방의','접수일자']).reset_index(drop=True)

    for idx, row in sorted_daily.iterrows():
        doc = row['처방의']; d = row['접수일자']; red = row['휴일']
        bg = ROW_A if ri % 2 == 0 else ROW_W

        dc(ws3, ri, 1, doc if doc!=prev_doc else '', bg=bg, align='left', bold=(doc!=prev_doc), fg=HDR_NAME)
        dc(ws3, ri, 2, str(d), bg=bg, fg=RED_FG if red else NRM_FG, align='left', bold=red)
        dc(ws3, ri, 3, row['요일'], bg=bg, fg=RED_FG if red else MUT_FG, bold=red)
        dc(ws3, ri, 4, row['주차'], bg=bg, fg=WK_FG, sz=9)
        dc(ws3, ri, 5, int(row['입원']), bg=bg)
        dc(ws3, ri, 6, int(row['외래']), bg=bg)
        dc(ws3, ri, 7, int(row['합계']), bg=bg, bold=True)
        ws3.row_dimensions[ri].height = 20

        nx = sorted_daily.iloc[idx+1] if idx+1 < len(sorted_daily) else None
        if nx is None or nx['처방의'] != doc:
            ri += 1
            d2 = sorted_daily[sorted_daily['처방의']==doc]
            bg2 = ROW_A if ri % 2 == 0 else ROW_W
            dc(ws3, ri, 1, f'{doc} 소계', bg=bg2, align='left', bold=True, fg=HDR_NAME, thick=True)
            for ci in [2,3,4]: dc(ws3, ri, ci, '', bg=bg2, thick=True)
            dc(ws3, ri, 5, int(d2['입원'].sum()), bg=bg2, bold=True, thick=True)
            dc(ws3, ri, 6, int(d2['외래'].sum()), bg=bg2, bold=True, thick=True)
            dc(ws3, ri, 7, int(d2['합계'].sum()), bg=bg2, bold=True, thick=True)
            ws3.row_dimensions[ri].height = 20
        prev_doc = doc; ri += 1

    ws3.column_dimensions['A'].width = 13; ws3.column_dimensions['B'].width = 13
    ws3.column_dimensions['C'].width = 6;  ws3.column_dimensions['D'].width = 12
    for col in ['E','F','G']: ws3.column_dimensions[col].width = 11

    # ── 시트4~: 처방의별 ───────────────────────────────────────────────
    for doc in docs:
        ws_d = wb.create_sheet(doc[:31])  # Excel 시트명 최대 31자
        ws_d.sheet_view.showGridLines = False
        title_cell(ws_d, 1, 6, f"{doc} — 일별 / 주별 집계")
        ws_d.row_dimensions[2].height = 6

        for ci, (h, bg) in enumerate([
            ('접수일자',HDR),('요일',HDR),('주차',WK_H),
            ('입원',IN_H),('외래',OUT_H),('합계',TOT_H)
        ], 1):
            hc(ws_d, 3, ci, h, bg=bg, align='left' if ci==1 else 'center')
        ws_d.row_dimensions[3].height = 26
        ws_d.freeze_panes = 'A4'

        doc_daily = daily[daily['처방의']==doc].sort_values('접수일자').reset_index(drop=True)
        ri = 4

        for idx, row in doc_daily.iterrows():
            d = row['접수일자']; red = row['휴일']
            bg = ROW_A if ri % 2 == 0 else ROW_W
            dc(ws_d, ri, 1, str(d), bg=bg, fg=RED_FG if red else NRM_FG, align='left', bold=red)
            dc(ws_d, ri, 2, row['요일'], bg=bg, fg=RED_FG if red else MUT_FG, bold=red)
            dc(ws_d, ri, 3, row['주차'], bg=bg, fg=WK_FG, sz=9)
            dc(ws_d, ri, 4, int(row['입원']), bg=bg)
            dc(ws_d, ri, 5, int(row['외래']), bg=bg)
            dc(ws_d, ri, 6, int(row['합계']), bg=bg, bold=True)
            ws_d.row_dimensions[ri].height = 20; ri += 1

            nx = doc_daily.iloc[idx+1] if idx+1 < len(doc_daily) else None
            if nx is None or nx['주차'] != row['주차']:
                wk = doc_daily[doc_daily['주차']==row['주차']]
                bg2 = ROW_A if ri % 2 == 0 else ROW_W
                dc(ws_d, ri, 1, f"{row['주차']} 소계", bg=bg2, align='left', bold=True, fg=HDR_NAME, thick=True)
                dc(ws_d, ri, 2, '', bg=bg2, thick=True); dc(ws_d, ri, 3, '', bg=bg2, thick=True)
                dc(ws_d, ri, 4, int(wk['입원'].sum()), bg=bg2, bold=True, thick=True)
                dc(ws_d, ri, 5, int(wk['외래'].sum()), bg=bg2, bold=True, thick=True)
                dc(ws_d, ri, 6, int(wk['합계'].sum()), bg=bg2, bold=True, thick=True)
                ws_d.row_dimensions[ri].height = 20; ri += 1

        hc(ws_d, ri, 1, '월 합계', align='left')
        dc(ws_d, ri, 2, '', bg=TOT_BG); dc(ws_d, ri, 3, '', bg=TOT_BG)
        dc(ws_d, ri, 4, int(doc_daily['입원'].sum()), bg=TOT_BG, bold=True)
        dc(ws_d, ri, 5, int(doc_daily['외래'].sum()), bg=TOT_BG, bold=True)
        dc(ws_d, ri, 6, int(doc_daily['합계'].sum()), bg=TOT_BG, bold=True)
        ws_d.row_dimensions[ri].height = 28

        ws_d.column_dimensions['A'].width = 13; ws_d.column_dimensions['B'].width = 6
        ws_d.column_dimensions['C'].width = 12
        for col in ['D','E','F']: ws_d.column_dimensions[col].width = 11

    # 정합성 검증 시트 추가 (있을 때만)
    if validation_result is not None:
        try:
            import sys, os
            sys.path.insert(0, os.path.dirname(__file__))
            from validator import add_validation_sheet
            add_validation_sheet(wb, validation_result)
        except Exception as e:
            print(f'[경고] 검증 시트 추가 실패: {e}')

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[Excel] 저장 완료: {output_path}")
    return output_path