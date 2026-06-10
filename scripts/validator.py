"""
validator.py — 주별 누적 vs 월별 데이터 정합성 검증
"""
from __future__ import annotations
import datetime
import pandas as pd
from typing import Dict, Any, List
from pathlib import Path


def validate(weekly_data: Dict[str, Any], monthly_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    주별 누적과 월별 집계 비교

    Args:
        weekly_data:  주별 파일 process() 결과 (또는 누적 주별)
        monthly_data: 월별 파일 process() 결과

    Returns:
        ok:       전체 일치 여부
        summary:  처방의별 입원/외래/합계 비교 행 목록
        detail:   날짜×처방의×구분 불일치 목록
        checked_at, weekly_period, monthly_period
    """
    w_sum = weekly_data['summary'].copy()
    m_sum = monthly_data['summary'].copy()
    w_sum['처방의'] = w_sum['처방의'].astype(str)
    m_sum['처방의'] = m_sum['처방의'].astype(str)

    present    = set(w_sum['처방의']) | set(m_sum['처방의'])
    order_keys = monthly_data.get('docs') or weekly_data.get('docs') or []
    all_docs   = [d for d in order_keys if d in present]
    all_docs  += sorted(present - set(all_docs))

    summary_rows: List[Dict] = []
    for doc in all_docs:
        for col in ('입원', '외래', '합계'):
            w_val = int(w_sum.loc[w_sum['처방의'] == doc, col].values[0]) \
                    if doc in w_sum['처방의'].values else 0
            m_val = int(m_sum.loc[m_sum['처방의'] == doc, col].values[0]) \
                    if doc in m_sum['처방의'].values else 0
            diff  = m_val - w_val
            summary_rows.append({
                '처방의':    doc,
                '구분':      col,
                '주별 누적': w_val,
                '월별 데이터': m_val,
                '차이':      diff,
                '상태':      '✓ 일치' if diff == 0 else f'⚠ 차이 {diff:+d}',
            })

    # 일별 상세 비교
    w_daily = weekly_data['daily'][['처방의','접수일자','입원','외래','합계']].copy()
    m_daily = monthly_data['daily'][['처방의','접수일자','입원','외래','합계']].copy()
    w_daily['처방의']   = w_daily['처방의'].astype(str)
    m_daily['처방의']   = m_daily['처방의'].astype(str)
    w_daily['접수일자'] = w_daily['접수일자'].astype(str)
    m_daily['접수일자'] = m_daily['접수일자'].astype(str)

    merged = pd.merge(
        w_daily.rename(columns={'입원':'주별_입원','외래':'주별_외래','합계':'주별_합계'}),
        m_daily.rename(columns={'입원':'월별_입원','외래':'월별_외래','합계':'월별_합계'}),
        on=['처방의', '접수일자'],
        how='outer',
    ).fillna(0)

    detail_rows: List[Dict] = []
    for _, row in merged.iterrows():
        for col in ('입원', '외래', '합계'):
            w_val = int(row[f'주별_{col}'])
            m_val = int(row[f'월별_{col}'])
            diff  = m_val - w_val
            if diff != 0:
                detail_rows.append({
                    '날짜':      row['접수일자'],
                    '처방의':    row['처방의'],
                    '구분':      col,
                    '주별 누적': w_val,
                    '월별 데이터': m_val,
                    '차이':      diff,
                })

    ok = bool(summary_rows) and all(r['차이'] == 0 for r in summary_rows)

    return {
        'ok':             ok,
        'summary':        summary_rows,
        'detail':         detail_rows,
        'checked_at':     datetime.datetime.now(),
        'weekly_period':  weekly_data['period'],
        'monthly_period': monthly_data['period'],
    }


def add_validation_sheet(wb, result: Dict[str, Any]) -> None:
    """openpyxl 워크북에 '정합성 검증' 시트 추가"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR    = '1E3A5F'; HDR_FG = 'FFFFFF'
    OK_BG  = 'E8F5E9'; OK_FG  = '1B5E20'
    ERR_BG = 'FFF3E0'; ERR_FG = 'E65100'
    DIFF_BG= 'FFF9C4'; ROW_A  = 'F2F6FC'
    BORD   = 'C5D0E8'

    def bs():
        s = Side(style='thin', color=BORD)
        return Border(left=s, right=s, top=s, bottom=s)

    def hc(ws, r, c, v, bg=HDR, fg=HDR_FG, bold=True, align='center'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='맑은 고딕', bold=bold, color=fg, size=10)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = bs()

    def dc(ws, r, c, v, bg='FFFFFF', fg='1A1A1A', bold=False, align='center'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='맑은 고딕', bold=bold, color=fg, size=10)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = bs()

    ok = result['ok']
    ws = wb.create_sheet('정합성 검증')
    ws.sheet_view.showGridLines = False

    # 헤더
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f"정합성 검증 — {'✓ 일치' if ok else '⚠ 불일치 있음'}"
    c.font  = Font(name='맑은 고딕', bold=True, size=14,
                   color=OK_FG if ok else ERR_FG)
    c.fill  = PatternFill('solid', start_color=OK_BG if ok else ERR_BG)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = bs()
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:F2')
    n = ws['A2']
    n.value = (f"검증: {result['checked_at'].strftime('%Y-%m-%d %H:%M')}  |  "
               f"주별: {result['weekly_period']}  |  월별: {result['monthly_period']}")
    n.font  = Font(name='맑은 고딕', size=9, color='666666', italic=True)
    n.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # 처방의별 요약
    ws.merge_cells('A4:F4')
    t = ws['A4']
    t.value = '▶ 처방의별 요약 비교'
    t.font  = Font(name='맑은 고딕', bold=True, size=11, color=HDR)
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[4].height = 24

    for ci, h in enumerate(('처방의','구분','주별 누적','월별 데이터','차이','상태'), 1):
        hc(ws, 5, ci, h, align='left' if ci <= 2 else 'center')
    ws.row_dimensions[5].height = 24

    for ri, row in enumerate(result['summary']):
        r = ri + 6
        is_diff = row['차이'] != 0
        bg = DIFF_BG if is_diff else (ROW_A if ri % 2 == 0 else 'FFFFFF')
        fg = ERR_FG  if is_diff else '1A1A1A'
        dc(ws, r, 1, row['처방의'],      bg=bg, fg=fg, bold=is_diff, align='left')
        dc(ws, r, 2, row['구분'],         bg=bg, fg=fg, align='left')
        dc(ws, r, 3, row['주별 누적'],    bg=bg, fg=fg)
        dc(ws, r, 4, row['월별 데이터'],  bg=bg, fg=fg)
        dc(ws, r, 5, row['차이'],         bg=bg, fg=fg, bold=is_diff)
        dc(ws, r, 6, row['상태'],         bg=bg,
           fg=OK_FG if not is_diff else ERR_FG, bold=True)
        ws.row_dimensions[r].height = 20

    sep = len(result['summary']) + 7
    ws.row_dimensions[sep - 1].height = 10

    # 일별 불일치 상세
    ws.merge_cells(f'A{sep}:F{sep}')
    t2 = ws.cell(row=sep, column=1)
    if result['detail']:
        t2.value = f"▶ 일별 불일치 상세 ({len(result['detail'])}건)"
        t2.font  = Font(name='맑은 고딕', bold=True, size=11, color=ERR_FG)
    else:
        t2.value = '▶ 일별 불일치 없음  ✓'
        t2.font  = Font(name='맑은 고딕', bold=True, size=11, color=OK_FG)
    t2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[sep].height = 24

    if result['detail']:
        hr = sep + 1
        for ci, h in enumerate(('날짜','처방의','구분','주별 누적','월별 데이터','차이'), 1):
            hc(ws, hr, ci, h, align='left' if ci <= 2 else 'center')
        ws.row_dimensions[hr].height = 24

        for ri, row in enumerate(result['detail']):
            r = hr + 1 + ri
            bg = DIFF_BG if ri % 2 == 0 else 'FFF8E1'
            dc(ws, r, 1, row['날짜'],          bg=bg, fg=ERR_FG, align='left')
            dc(ws, r, 2, row['처방의'],         bg=bg, fg=ERR_FG, align='left', bold=True)
            dc(ws, r, 3, row['구분'],            bg=bg)
            dc(ws, r, 4, row['주별 누적'],       bg=bg)
            dc(ws, r, 5, row['월별 데이터'],     bg=bg)
            dc(ws, r, 6, row['차이'],            bg=bg, fg=ERR_FG, bold=True)
            ws.row_dimensions[r].height = 20

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
