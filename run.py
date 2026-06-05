"""
run.py — 물리치료 집계 자동화 메인
Python 3.8+ 호환

사용법:
  python run.py                  # 오늘 날짜 기준 자동 실행
  python run.py 파일.xlsx        # 특정 파일 처리
  python run.py --force          # 전체 강제 실행 (날짜/중복 무시)
  python run.py --weekly         # 주별만 즉시 실행 (GUI 버튼용)
  python run.py --monthly        # 월별만 즉시 실행 (GUI 버튼용)
  python run.py --validate       # 정합성 검증만
  python run.py --schedule       # 스케줄러 데몬 모드 (창 켜놔야 함)
  ※ 창 없는 자동 실행은 Windows 작업 스케줄러 + run_silent.vbs 사용
"""
from __future__ import annotations

import sys
import logging
import logging.handlers
import datetime
import traceback
import time
from pathlib import Path
from typing import Optional, Set, Tuple

for _stream in (sys.stdout, sys.stderr):
    if _stream and hasattr(_stream, 'reconfigure'):
        try:
            _stream.reconfigure(encoding='utf-8')
        except Exception:
            pass

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / 'scripts'))

DATA_WEEKLY   = ROOT / 'data' / 'weekly'
DATA_MONTHLY  = ROOT / 'data' / 'monthly'
PROCESSED_LOG = ROOT / 'processed.log'
LOG_FILE      = ROOT / 'run.log'

# ── 로깅 (10MB 로테이션, 3개 보관) ──────────────────────────────────────
def _setup_logging() -> logging.Logger:
    fmt = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    logger = logging.getLogger('pt_report')
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)

    if sys.stdout is not None:
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(fmt)
        logger.addHandler(ch)

    try:
        fh = logging.handlers.RotatingFileHandler(
            LOG_FILE, maxBytes=10*1024*1024, backupCount=3, encoding='utf-8'
        )
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    except Exception:
        pass
    return logger

log = _setup_logging()

# ── 지연 임포트 ───────────────────────────────────────────────────────────
def _get_process():
    from process import load_and_process, get_file_year_month, load_and_merge
    return load_and_process, get_file_year_month, load_and_merge

def _get_generators():
    from excel_report import generate_excel
    from html_dashboard import generate_html
    return generate_excel, generate_html

# ── output 경로 ───────────────────────────────────────────────────────────
def get_output_dirs(year: int, month: int) -> Tuple[Path, Path]:
    ym = f'{year}-{month:02d}'
    reports   = ROOT / 'output' / ym / 'reports'
    dashboard = ROOT / 'output' / ym / 'dashboard'
    reports.mkdir(parents=True, exist_ok=True)
    dashboard.mkdir(parents=True, exist_ok=True)
    return reports, dashboard

def _move_to_done(filepath: Path) -> None:
    """처리 완료 파일을 data/_done/YYYY-MM/ 으로 이동"""
    try:
        import shutil
        from process import get_file_year_month
        ym = get_file_year_month(filepath)
        if ym:
            done_dir = filepath.parent / '_done' / f'{ym[0]}-{ym[1]:02d}'
        else:
            done_dir = filepath.parent / '_done'
        done_dir.mkdir(parents=True, exist_ok=True)
        dest = done_dir / filepath.name
        if dest.exists():
            if dest.stat().st_size == filepath.stat().st_size:
                # 동일 파일이 이미 _done에 존재 → 소스 삭제 (타임스탬프 복사본 생성 안 함)
                filepath.unlink()
                log.info(f"중복 파일 제거 (이미 _done에 존재): {filepath.name}")
                return
            stamp = datetime.datetime.now().strftime('%H%M%S')
            dest  = done_dir / f"{filepath.stem}_{stamp}{filepath.suffix}"
        shutil.move(str(filepath), str(dest))
        log.info(f"원본 파일 이동: {filepath.name} → data/_done/")
    except Exception as e:
        log.warning(f"원본 파일 이동 실패 (파일은 그대로): {e}")


def _cleanup_old_outputs(reports_dir: Path, dashboard_dir: Path) -> None:
    """같은 기간 이전 결과물을 _archive/ 로 이동 (최신 1개만 유지)"""
    archive_r = reports_dir   / '_archive'
    archive_d = dashboard_dir / '_archive'

    for folder, archive, suffix in [
        (reports_dir,   archive_r, '.xlsx'),
        (dashboard_dir, archive_d, '.html'),
    ]:
        files = sorted(
            [f for f in folder.glob(f'*{suffix}') if not f.name.startswith('정합성')],
            key=lambda p: p.stat().st_mtime
        )
        if len(files) > 1:
            archive.mkdir(exist_ok=True)
            for old_file in files[:-1]:
                try:
                    old_file.rename(archive / old_file.name)
                except Exception:
                    pass

# ── 처리 완료 기록 ────────────────────────────────────────────────────────
def _load_processed() -> Set[str]:
    """오늘 처리 완료한 '파일경로|수정시간' 집합 반환"""
    if not PROCESSED_LOG.exists():
        return set()
    today = datetime.date.today().isoformat()
    result = set()
    try:
        for line in PROCESSED_LOG.read_text(encoding='utf-8').splitlines():
            parts = line.strip().split('|')
            if len(parts) >= 2 and parts[0] == today:
                key = '|'.join(parts[1:])
                result.add(key)
    except Exception:
        pass
    return result

def _is_already_processed(filepath: Path, processed: Set[str]) -> bool:
    try:
        mtime = int(filepath.stat().st_mtime)
        key = f"{filepath.resolve()}|{mtime}"
        if key in processed:
            return True
    except Exception:
        pass
    return False

def _mark_processed(filepath: Path) -> None:
    today = datetime.date.today().isoformat()
    try:
        mtime = int(filepath.stat().st_mtime)
        with open(PROCESSED_LOG, 'a', encoding='utf-8') as f:
            f.write(f"{today}|{filepath.resolve()}|{mtime}\n")
    except Exception as e:
        log.warning(f"처리 기록 저장 실패: {e}")

def _cleanup_processed_log() -> None:
    """processed.log 에서 30일 초과 항목 제거"""
    if not PROCESSED_LOG.exists():
        return
    cutoff = (datetime.date.today() - datetime.timedelta(days=30)).isoformat()
    try:
        lines = PROCESSED_LOG.read_text(encoding='utf-8').splitlines()
        kept  = [l for l in lines if l.strip() and l.strip().split('|')[0] >= cutoff]
        if len(kept) < len(lines):
            PROCESSED_LOG.write_text(
                '\n'.join(kept) + ('\n' if kept else ''), encoding='utf-8'
            )
    except Exception:
        pass

# ── 파일 유효성 ───────────────────────────────────────────────────────────
def is_valid_excel(filepath: Path) -> bool:
    if not filepath.is_file():
        return False
    ext = filepath.suffix.lower()
    if ext not in ('.xlsx', '.xls'):
        return False
    if filepath.stat().st_size < 1024:
        return False
    try:
        if ext == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            wb.close()
        else:
            import xlrd
            xlrd.open_workbook(str(filepath))
        return True
    except Exception:
        return False

# ── 파일 탐색 ─────────────────────────────────────────────────────────────
def _glob_excel(folder: Path) -> list:
    """xls/xlsx (대소문자 무관) 탐색 (수정시간 내림차순)"""
    if not folder.exists():
        return []
    files = [f for f in folder.iterdir()
             if f.is_file() and f.suffix.lower() in ('.xlsx', '.xls')]
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)

def _find_latest_for_month(
    folder: Path, year: int, month: int, processed: Set[str]
) -> Optional[Path]:
    _, get_ym, _ = _get_process()
    candidates = [
        f for f in _glob_excel(folder)
        if is_valid_excel(f)
        and not _is_already_processed(f, processed)
        and get_ym(f) == (year, month)
    ]
    return candidates[0] if candidates else None

def _find_any_for_month(folder: Path, year: int, month: int) -> Optional[Path]:
    _, get_ym, _ = _get_process()
    candidates = [
        f for f in _glob_excel(folder)
        if is_valid_excel(f) and get_ym(f) == (year, month)
    ]
    return candidates[0] if candidates else None

def _find_all_year_months() -> list:
    _, get_ym, _ = _get_process()
    ym_set = set()
    for folder in (DATA_WEEKLY, DATA_MONTHLY):
        for f in _glob_excel(folder):
            if is_valid_excel(f):
                ym = get_ym(f)
                if ym:
                    ym_set.add(ym)
    return sorted(ym_set)

def _base_stem(stem: str) -> str:
    """'파일명_235542' → '파일명'  (_HHMMSS 충돌 접미사 제거)"""
    if len(stem) > 7 and stem[-7] == '_' and stem[-6:].isdigit():
        return stem[:-7]
    return stem

def _collect_weekly_files(year: int, month: int) -> list:
    """
    data/weekly/ + _done/YYYY-MM/ 에서 같은 달 파일 전부 수집 (오래된 순).
    _HHMMSS 타임스탬프 중복 파일은 최신 1개만 유지.
    """
    _, get_ym, _ = _get_process()
    all_files = []
    for search_dir in [DATA_WEEKLY, DATA_WEEKLY / '_done' / f'{year}-{month:02d}']:
        for f in _glob_excel(search_dir):  # 폴더 없으면 빈 list 반환
            if is_valid_excel(f) and get_ym(f) == (year, month):
                all_files.append(f)
    all_files.sort(key=lambda p: p.stat().st_mtime)

    # 베이스명 기준 중복 제거 — 최신 파일만 유지
    stem_to_file: dict = {}
    for f in all_files:
        base = _base_stem(f.stem).lower()
        stem_to_file[base] = f  # 오래된 순 순회하므로 마지막(최신)이 남음
    return sorted(stem_to_file.values(), key=lambda p: p.stat().st_mtime)

# ── 날짜 판단 ─────────────────────────────────────────────────────────────
def _get_today_mode(today: datetime.date) -> Optional[str]:
    if today.weekday() == 6:
        log.info(f"일요일({today}) — 스킵")
        return None
    if today.day in (1, 2):
        return 'monthly'
    if today.weekday() in (0, 1):
        return 'weekly'
    log.info(f"실행 대상 날짜 아님({today}) — 스킵")
    return None

def _get_target_ym(today: datetime.date, mode: str) -> Tuple[int, int]:
    if mode == 'monthly':
        prev = today.replace(day=1) - datetime.timedelta(days=1)
        return prev.year, prev.month
    return today.year, today.month

# ── 출력 생성 (데이터 → Excel + HTML) ────────────────────────────────────
def _save_outputs(
    data: dict,
    counterpart: Optional[Path] = None,
    counterpart_data: Optional[dict] = None,
) -> bool:
    """처리된 데이터를 받아 Excel + HTML 생성 및 archive 정리"""
    generate_excel, generate_html = _get_generators()

    for w in data.get('warnings', []):
        log.warning(w)

    year, month = data['year_month']
    period      = data['period']
    stamp       = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

    reports_dir, dashboard_dir = get_output_dirs(year, month)
    excel_path = reports_dir   / f"집계리포트_{period}_{stamp}.xlsx"
    html_path  = dashboard_dir / f"대시보드_{period}_{stamp}.html"

    # 정합성 검증
    # counterpart_data: 이미 로드/합산된 dict (월별 검증 시 주별 합산 데이터)
    # counterpart: 단일 파일 Path (주별 검증 시 월별 파일)
    validation_result = None
    other_data = counterpart_data  # 미리 전달된 경우 우선 사용
    if other_data is None and counterpart and is_valid_excel(counterpart):
        try:
            load_fn, _, _ = _get_process()
            other_data = load_fn(counterpart)
        except Exception as e:
            log.warning(f"정합성 검증 실패 (리포트는 생성 계속): {e}")

    if other_data is not None:
        try:
            if other_data['year_month'] == (year, month):
                from validator import validate
                # counterpart_data 전달 시 other=주별, data=월별
                is_monthly_counterpart = (counterpart is not None and
                                          counterpart.parent == DATA_MONTHLY)
                if is_monthly_counterpart:
                    validation_result = validate(data, other_data)
                else:
                    validation_result = validate(other_data, data)
                status = '✓ 일치' if validation_result['ok'] else \
                         f"⚠ 불일치 {len(validation_result['detail'])}건"
                log.info(f"정합성 검증: {status}")
            else:
                log.info("정합성 검증 스킵 (연월 불일치)")
        except Exception as e:
            log.warning(f"정합성 검증 실패 (리포트는 생성 계속): {e}")

    try:
        generate_excel(data, excel_path, validation_result=validation_result)
        generate_html(data, html_path)
    except Exception as e:
        log.error(f"출력 생성 실패: {e}")
        log.debug(traceback.format_exc())
        return False

    _cleanup_old_outputs(reports_dir, dashboard_dir)
    log.info(f"✓ 완료: {period} → output/{year}-{month:02d}/")
    return True

# ── 단일 파일 처리 (재시도 포함) ─────────────────────────────────────────
def process_file(
    filepath: Path,
    mark: bool = True,
    counterpart: Optional[Path] = None,
    counterpart_data: Optional[dict] = None,
    retry: bool = True,
) -> bool:
    """
    filepath:         처리할 파일
    mark:             완료 기록 여부
    counterpart:      정합성 비교 단일 파일 (주별 실행 시 월별 파일)
    counterpart_data: 정합성 비교 pre-loaded dict (월별 실행 시 주별 합산 데이터)
    retry:            파일 잠김 시 30분 후 1회 재시도
    """
    load_and_process, _, _ = _get_process()
    log.info(f"처리 시작: {filepath.name}")

    for attempt in range(2):
        try:
            data = load_and_process(filepath)
            break
        except PermissionError as e:
            if attempt == 0 and retry:
                log.warning(f"파일 잠김 — 30분 후 재시도: {filepath.name}")
                time.sleep(30 * 60)
                continue
            log.error(str(e))
            return False
        except ValueError as e:
            log.error(f"데이터 오류: {e}")
            return False
        except Exception as e:
            log.error(f"처리 실패: {filepath.name} — {e}")
            log.debug(traceback.format_exc())
            return False
    else:
        return False

    ok = _save_outputs(data, counterpart=counterpart, counterpart_data=counterpart_data)
    if ok and mark:
        _mark_processed(filepath)
        _move_to_done(filepath)
    return ok

# ── 주별 / 월별 실행 ──────────────────────────────────────────────────────
def _run_weekly(year: int, month: int, processed: Set[str], mark: bool = True) -> None:
    log.info(f"=== 주별 집계 {year}-{month:02d} ===")

    new_fp = _find_latest_for_month(DATA_WEEKLY, year, month, processed)
    if not new_fp:
        log.info("새로 처리할 파일 없음 (없거나 이미 처리됨)")
        return

    all_files = _collect_weekly_files(year, month)
    if not all_files:
        all_files = [new_fp]

    counterpart = _find_any_for_month(DATA_MONTHLY, year, month)
    load_and_process, get_ym, load_and_merge = _get_process()

    log.info(f"주별 파일 {len(all_files)}개 합산: {[f.name for f in all_files]}"
             if len(all_files) > 1 else f"주별 파일: {all_files[0].name}")

    for attempt in range(2):
        try:
            data = load_and_merge(all_files) if len(all_files) > 1 \
                   else load_and_process(all_files[0])
            break
        except PermissionError as e:
            if attempt == 0:
                log.warning(f"파일 잠김 — 30분 후 재시도")
                time.sleep(30 * 60)
                continue
            log.error(str(e))
            return
        except ValueError as e:
            log.error(f"데이터 오류: {e}")
            return
        except Exception as e:
            log.error(f"처리 실패: {e}")
            log.debug(traceback.format_exc())
            return
    else:
        return

    ok = _save_outputs(data, counterpart=counterpart)
    if ok and mark:
        # all_files 기반 대신 DATA_WEEKLY 직접 스캔:
        # mtime 역전 등으로 all_files에서 누락된 파일도 빠짐없이 이동
        for f in list(_glob_excel(DATA_WEEKLY)):
            ym = get_ym(f)
            if ym and ym == (year, month):
                _mark_processed(f)
                _move_to_done(f)

def _run_monthly(year: int, month: int, processed: Set[str], mark: bool = True) -> None:
    log.info(f"=== 월별 집계 {year}-{month:02d} ===")
    fp = _find_latest_for_month(DATA_MONTHLY, year, month, processed)
    if not fp:
        log.info("data/monthly/ 파일 없음 → data/weekly/ 최신 파일로 대체")
        fp = _find_latest_for_month(DATA_WEEKLY, year, month, processed)
    if not fp:
        log.info("처리할 파일 없음")
        return

    # 정합성 검증용 주별 합산 데이터 수집 (weekly/ + _done/ 전체)
    counterpart_data = None
    if fp.parent == DATA_MONTHLY:
        load_and_process, get_ym, load_merge = _get_process()
        all_weekly = _collect_weekly_files(year, month)
        if all_weekly:
            try:
                counterpart_data = load_merge(all_weekly) \
                                   if len(all_weekly) > 1 \
                                   else load_and_process(all_weekly[0])
                log.info(f"정합성 검증용 주별 파일 {len(all_weekly)}개 합산")
            except Exception as e:
                log.warning(f"주별 데이터 로드 실패 — 정합성 검증 스킵: {e}")

    process_file(fp, mark=mark, counterpart_data=counterpart_data)

# ── 자동 실행 ─────────────────────────────────────────────────────────────
def run_auto(force: bool = False) -> None:
    today     = datetime.date.today()
    processed = set() if force else _load_processed()

    if force:
        ym_list = _find_all_year_months()
        if not ym_list:
            log.info("처리할 파일 없음")
            return
        log.info(f"강제 실행: {[f'{y}-{m:02d}' for y, m in ym_list]}")
        for year, month in ym_list:
            _run_weekly(year, month, processed, mark=False)
            _run_monthly(year, month, processed, mark=False)
        return

    mode = _get_today_mode(today)
    if mode is None:
        return

    year, month = _get_target_ym(today, mode)
    if mode == 'monthly':
        _run_monthly(year, month, processed)
    else:
        _run_weekly(year, month, processed)

# ── 정합성 검증 단독 ──────────────────────────────────────────────────────
def run_validate_only() -> None:
    today = datetime.date.today()
    mode  = 'monthly' if today.day in (1, 2) else 'weekly'
    year, month = _get_target_ym(today, mode)

    log.info(f"=== 정합성 검증 {year}-{month:02d} ===")

    weekly_fp  = _find_any_for_month(DATA_WEEKLY,  year, month)
    monthly_fp = _find_any_for_month(DATA_MONTHLY, year, month)

    if not weekly_fp:
        log.warning(f"data/weekly/ 에 {year}-{month:02d} 파일 없음")
        return
    if not monthly_fp:
        log.warning(f"data/monthly/ 에 {year}-{month:02d} 파일 없음")
        return

    try:
        load_and_process, _, _ = _get_process()
        from validator import validate, add_validation_sheet
        import openpyxl

        w_data = load_and_process(weekly_fp)
        m_data = load_and_process(monthly_fp)

        for w in w_data.get('warnings', []) + m_data.get('warnings', []):
            log.warning(w)

        result = validate(w_data, m_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '커버'
        ws['A1'] = (f"정합성 검증 — {year}-{month:02d}  "
                    f"({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})")
        add_validation_sheet(wb, result)

        reports_dir, _ = get_output_dirs(year, month)
        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        out = reports_dir / f"정합성검증_{year}{month:02d}_{stamp}.xlsx"
        wb.save(out)

        status = '✓ 일치' if result['ok'] else f"⚠ 불일치 {len(result['detail'])}건"
        log.info(f"검증 완료: {status} → {out.name}")

    except Exception as e:
        log.error(f"검증 실패: {e}")
        log.debug(traceback.format_exc())

# ── 스케줄러 데몬 ─────────────────────────────────────────────────────────
def run_schedule_mode() -> None:
    try:
        from apscheduler.schedulers.blocking import BlockingScheduler
        from apscheduler.triggers.cron import CronTrigger
    except ImportError:
        log.error("apscheduler 미설치:\n  pip install apscheduler")
        sys.exit(1)

    scheduler = BlockingScheduler(timezone='Asia/Seoul')
    scheduler.add_job(run_auto, CronTrigger(day_of_week='mon', hour=12),
                      id='weekly', misfire_grace_time=600, coalesce=True)
    scheduler.add_job(run_auto, CronTrigger(day='1', hour=12),
                      id='monthly', misfire_grace_time=600, coalesce=True)

    log.info("스케줄러 시작 — 주별 월요일 12:00 / 월별 1일 12:00")
    log.info("※ 창 없이 실행하려면 이 모드 대신 Windows 작업 스케줄러 사용")
    try:
        scheduler.start()
    except (KeyboardInterrupt, SystemExit):
        log.info("스케줄러 종료")

# ── 메인 ──────────────────────────────────────────────────────────────────
def main() -> None:
    DATA_WEEKLY.mkdir(parents=True, exist_ok=True)
    DATA_MONTHLY.mkdir(parents=True, exist_ok=True)
    _cleanup_processed_log()

    args = sys.argv[1:]

    if '--schedule' in args:
        run_schedule_mode()
    elif '--validate' in args:
        run_validate_only()
    elif '--weekly' in args:
        today = datetime.date.today()
        ym_list = _find_all_year_months()
        year, month = ym_list[-1] if ym_list else (today.year, today.month)
        _run_weekly(year, month, set(), mark=True)
    elif '--monthly' in args:
        today = datetime.date.today()
        ym_list = _find_all_year_months()
        year, month = ym_list[-1] if ym_list else (today.year, today.month)
        _run_monthly(year, month, set(), mark=True)
    elif '--force' in args:
        remaining = [a for a in args if a != '--force']
        if remaining:
            fp = Path(remaining[0])
            if not is_valid_excel(fp):
                log.error(f"유효하지 않은 파일: {fp}")
                sys.exit(1)
            process_file(fp, mark=False)
        else:
            run_auto(force=True)
    elif args and not args[0].startswith('--'):
        fp = Path(args[0])
        if not is_valid_excel(fp):
            log.error(f"유효하지 않은 파일: {fp}")
            sys.exit(1)
        process_file(fp)
    else:
        run_auto()

if __name__ == '__main__':
    main()
